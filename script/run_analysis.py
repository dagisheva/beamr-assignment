import subprocess
import os
import re
import sys
import shutil
import statistics
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INPUT_FILE = "foreman-cif.yuv"
RESOLUTION = "352x288"
OUTPUT_DIR = "encoded_files"
RESULTS_FILE = "qp_analysis_report.xlsx"
QP_RANGE = range(1, 52)
RUNS_PER_QP = 3


def run_x264_once(qp, input_file, resolution, output_file):
    cmd = ["x264", "--qp", str(qp), "--input-res", resolution, "-o", output_file, input_file]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
        if result.returncode != 0:
            print(f"\n    x264 exited with code {result.returncode}", end="", flush=True)
            return None
        output = result.stderr + result.stdout
        m = re.search(r'encoded \d+ frames, (\d+\.\d+) fps', output)
        return float(m.group(1)) if m else None
    except subprocess.TimeoutExpired:
        print("\n    run timed out", end="", flush=True)
        return None
    except OSError as e:
        print(f"\n    OS error: {e}", end="", flush=True)
        return None


def run_x264(qp, input_file, resolution, output_dir, runs):
    output_file = os.path.join(output_dir, f"output_qp{qp}.264")
    fps_samples = []
    for _ in range(runs):
        fps = run_x264_once(qp, input_file, resolution, output_file)
        if fps is not None:
            fps_samples.append(fps)
    file_size = os.path.getsize(output_file) if os.path.exists(output_file) else None
    median_fps = statistics.median(fps_samples) if fps_samples else None
    return {
        "qp": qp,
        "file_size_kb": file_size / 1024 if file_size is not None else None,
        "median_fps": median_fps,
    }


def add_line_chart(ws, title, y_axis_label, data_col, last_row, anchor):
    chart = LineChart()
    chart.title = title
    chart.x_axis.title = "QP"
    chart.y_axis.title = y_axis_label
    chart.style = 10
    chart.width = 20
    chart.height = 12
    chart.add_data(Reference(ws, min_col=data_col, min_row=1, max_row=last_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last_row))
    chart.series[0].graphicalProperties.line.width = 25000
    ws.add_chart(chart, anchor)


def write_xlsx(results, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "QP Sweep Results"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    alt_fill = PatternFill("solid", fgColor="D6E4F0")

    headers = ["QP", "File Size (KB)", "FPS (median)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border

    for i, r in enumerate(results):
        row = i + 2
        ws.cell(row=row, column=1, value=r["qp"]).font = data_font
        ws.cell(row=row, column=2, value=r["file_size_kb"]).font = data_font
        ws.cell(row=row, column=2).number_format = '#,##0.0'
        ws.cell(row=row, column=3, value=r["median_fps"]).font = data_font
        if r["median_fps"] is not None:
            ws.cell(row=row, column=3).number_format = '0.00'
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
            if i % 2 == 1:
                ws.cell(row=row, column=col).fill = alt_fill

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16

    last_row = len(results) + 1
    add_line_chart(ws, "File Size vs QP", "File Size (KB)", 2, last_row, "F2")
    add_line_chart(ws, "Encoding Speed (FPS) vs QP", "FPS", 3, last_row, "F23")

    wb.save(filename)


def main():
    if not shutil.which("x264"):
        print("Error: x264 not found on PATH.")
        print("Install it (e.g. 'sudo apt install x264' or 'brew install x264') and try again.")
        sys.exit(1)

    if not os.path.exists(INPUT_FILE):
        print(f"Error: {INPUT_FILE} not found in current directory.")
        print("Place the YUV file in the same directory and re-run.")
        sys.exit(1)

    total_encodes = len(QP_RANGE) * RUNS_PER_QP
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"Starting x264 QP sweep (QP 1-51, {RUNS_PER_QP} runs each = {total_encodes} encodes) on {INPUT_FILE}...")
    results = []
    for qp in QP_RANGE:
        print(f"  QP={qp:2d} ({RUNS_PER_QP} runs)...", end=" ", flush=True)
        r = run_x264(qp, INPUT_FILE, RESOLUTION, OUTPUT_DIR, RUNS_PER_QP)
        size_str = f"{r['file_size_kb']:8.1f} KB" if r["file_size_kb"] is not None else "     N/A   "
        fps_str = f"{r['median_fps']:6.2f}" if r["median_fps"] is not None else "N/A"
        print(f"Size: {size_str} | FPS: {fps_str}")
        results.append(r)

    write_xlsx(results, RESULTS_FILE)
    print(f"\nResults saved to {RESULTS_FILE}")


if __name__ == "__main__":
    main()